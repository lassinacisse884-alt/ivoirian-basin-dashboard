#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
update_dashboard.py
-------------------
Lit le fichier Excel IVOIRIAN_SEDIMENTARY_BASIN_DATA_BASE.xlsm et met a jour
automatiquement le dashboard HTML (IVOIRIAN_BASIN_DASHBOARD.html).

Usage:
  python update_dashboard.py          -> mise a jour unique
  python update_dashboard.py --watch  -> surveille le fichier et met a jour automatiquement

Dependances: openpyxl, watchdog (pip install openpyxl watchdog)
"""
import sys
import os
import json
import re
import time
import argparse
from datetime import datetime, date, timedelta
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string as c2i

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────
EXCEL_PATH = r"D:\LCISSE\Desktop\IVOIRIAN SEDIMENTARY BASIN OIL AND GAS\1-GENERAL DATABASE\IVOIRIAN_SEDIMENTARY_BASIN_DATA_BASE.xlsm"
DASHBOARD_PATH = r"D:\LCISSE\Desktop\IVOIRIAN SEDIMENTARY BASIN OIL AND GAS\1-GENERAL DATABASE\IVOIRIAN_BASIN_DASHBOARD.html"

# ─────────────────────────────────────────────
# UTILITAIRES
# ─────────────────────────────────────────────
NULL_VALUES = {None, "N/A", "NA", "-", "NT", "S/I", "F/S", "Faulty", "Traces", "L.P.", "LP", ""}

def clean(v):
    """Convertit les valeurs non-exploitables en None, arrondit les flottants."""
    if v is None:
        return None
    if isinstance(v, str):
        # Supprimer les caracteres de controle (retours a la ligne dans les cellules Excel)
        vs = re.sub(r'[\x00-\x1f\x7f]', ' ', v).strip()
        if vs in NULL_VALUES or vs == "":
            return None
        # Texte valide (statut de puits, etc.)
        return vs
    if isinstance(v, timedelta):
        # Convertir timedelta en heures (ex: "1 day, 0:00:00" -> 24.0)
        return round(v.total_seconds() / 3600, 4)
    if isinstance(v, (datetime, date)):
        if isinstance(v, datetime):
            return v.date().isoformat()
        return v.isoformat()
    if isinstance(v, float):
        if v != v:  # NaN
            return None
        return round(v, 6)
    if isinstance(v, int):
        return v
    return str(v)

def fmt_date(v):
    """Retourne une date ISO ou None."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, str):
        try:
            return datetime.strptime(v.strip(), "%Y-%m-%d").date().isoformat()
        except:
            return None
    return None

def rows(ws, start_row, end_col=None):
    """Generateur de lignes depuis start_row; s'arrete a la premiere ligne vide."""
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        # S'arreter si toute la ligne est vide
        non_null = [v for v in row if v is not None]
        if not non_null:
            break
        yield row

def get_col(row, col_letter):
    """Lit la valeur d'une colonne (lettre) dans une ligne (tuple 0-indexed)."""
    idx = c2i(col_letter) - 1
    if idx < len(row):
        return row[idx]
    return None

def extract_simple_section(ws, start_row, date_col, field_map, max_rows=1500):
    """
    Extrait une section simple (1 ligne = 1 date).
    field_map = {json_key: 'COLONNE_LETTRE', ...}
    """
    result = []
    count = 0
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        d = fmt_date(get_col(row, date_col))
        if d is None:
            break
        record = {"date": d}
        for key, col in field_map.items():
            record[key] = clean(get_col(row, col))
        result.append(record)
        count += 1
        if count >= max_rows:
            break
    return result

def extract_multirow_section(ws, start_row, date_col, id_col, field_map, max_rows=20000):
    """
    Extrait une section avec plusieurs lignes par date (wells, chemicals, energy).
    date_col: colonne de date (peut etre repete ou vide)
    id_col: colonne identifiant (well_id, label)
    """
    result = []
    current_date = None
    count = 0
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        d_raw = get_col(row, date_col)
        id_raw = get_col(row, id_col)
        # La date peut etre presente seulement sur la 1ere ligne du groupe
        if d_raw is not None:
            d = fmt_date(d_raw)
            if d:
                current_date = d
        if current_date is None:
            continue
        if id_raw is None or (isinstance(id_raw, str) and id_raw.strip() in NULL_VALUES):
            continue
        record = {"date": current_date}
        for key, col in field_map.items():
            record[key] = clean(get_col(row, col))
        result.append(record)
        count += 1
        if count >= max_rows:
            break
    return result

# ─────────────────────────────────────────────
# EXTRACTION PAR FEUILLE
# ─────────────────────────────────────────────

def extract_bp1(ws):
    """BALEINE PHASE 1"""
    data = {}

    # OIL PRODUCTION DATA (col E-L)
    data["oil"] = extract_simple_section(ws, 11, "E", {
        "gross": "F", "net": "G", "salt_content": "H", "api": "I",
        "bsw": "J", "rvp": "K", "delta": "L"
    })

    # GAS PRODUCTION DATA (col N-Y)
    data["gas"] = extract_simple_section(ws, 11, "N", {
        "gas_prod": "O", "gas_export": "P", "fuel_gas": "Q",
        "hc_dew": "R", "h2o_dew": "S", "h2s": "T",
        "wobbe": "U", "cal_value": "V", "co2": "W", "flare": "X", "delta": "Y"
    })

    # WATER PRODUCTION DATA (col AA-AK)
    data["water"] = extract_simple_section(ws, 11, "AA", {
        "dilution": "AB", "prod_overboard": "AC", "prod_slop": "AD",
        "test_sep": "AE", "water_produced": "AF", "oil_in_water": "AG",
        "o2": "AH", "chlorine": "AI", "sulphate": "AJ", "delta": "AK"
    })

    # PRODUCER WELLS DATA (col AM-BG, multiple rows per date)
    data["wells"] = extract_multirow_section(ws, 11, "AM", "AN", {
        "well_id": "AN", "status": "AO", "flowing_hrs": "AP", "choke": "AQ",
        "whp": "AR", "annulus": "AS", "wht": "AT", "dsch_p": "AU", "dsch_t": "AV",
        "bht": "AW", "bhp": "AX", "oil_rate": "AY", "gas_rate": "AZ",
        "water_rate": "BA", "gor": "BB", "wlr": "BC", "glr": "BD"
    })

    # OIL STOCK DATA (col BI-BN)
    data["stocks"] = extract_simple_section(ws, 11, "BI", {
        "gross": "BJ", "net": "BK", "dead": "BL", "pumpable": "BM", "days_full": "BN"
    })

    # HSE DATA (col BP-BY)
    data["hse"] = extract_simple_section(ws, 11, "BP", {
        "days_lti": "BQ", "mtc": "BR", "fac": "BS", "near_miss": "BT",
        "oil_spill_sea": "BU", "chem_spill": "BV", "hoc": "BW", "oil_spill": "BX", "drill": "BY"
    })

    # SUBSEA CHEMICAL DATA (col CA-CH, label rows)
    data["chem_subsea"] = extract_multirow_section(ws, 11, "CA", "CB", {
        "label": "CB", "meoh_fwd": "CC", "ldhi": "CD", "wi_ppd": "CE",
        "lci": "CF", "scale_inh": "CG", "subsea_ci": "CH"
    })

    # TOPSIDE CHEMICAL DATA (col CJ-CZ, label rows)
    data["chem_topside"] = extract_multirow_section(ws, 11, "CJ", "CK", {
        "label": "CK", "meoh_aft": "CL", "wi_ppd": "CM", "teg": "CN",
        "eb": "CO", "h2so4": "CP", "lci": "CQ", "scale_inh": "CR",
        "poly_m16": "CS", "gas_ci": "CT", "ultimer_m23": "CU", "naoh": "CV",
        "teg_ci": "CW", "teg_ph": "CX", "teg_antifoam": "CY", "vaptreat": "CZ"
    })

    # ENERGY PRODUCTION DATA (col DB-DI, label rows)
    data["energy"] = extract_multirow_section(ws, 11, "DB", "DC", {
        "label": "DC", "gtg1": "DD", "gtg2": "DE", "edg1": "DF",
        "edg2": "DG", "boiler_p": "DH", "boiler_s": "DI"
    })

    # LIQUID FUEL STORAGE DATA (col DK-DO)
    data["fuel"] = extract_simple_section(ws, 11, "DK", {
        "mgo": "DL", "pres_diesel": "DM", "bunkering_mgo": "DN", "bunkering_pres": "DO"
    })

    return data


def extract_bp2(ws):
    """BALEINE PHASE 2"""
    data = {}

    # Field Daily Oil Production (E-I)
    data["oil"] = extract_simple_section(ws, 11, "E", {
        "gross": "F", "net": "G", "bsw": "H", "delta": "I"
    })

    # Field Daily Gas Production (K-Q)
    data["gas"] = extract_simple_section(ws, 11, "K", {
        "gas_prod": "L", "gas_export": "M", "fuel_gas": "N",
        "gor": "O", "gas_lift": "P", "delta": "Q"
    })

    # Daily Water Production (S-AA)
    data["water"] = extract_simple_section(ws, 11, "S", {
        "water_prod": "T", "prod_overboard": "U", "oil_in_water": "V",
        "water_reinj": "W", "pw_reinj": "X", "sw_inj": "Y",
        "o2": "Z", "wi_pressure": "AA"
    })

    # Oil Production Wells (AC-AS, multiple per date)
    data["prod_wells"] = extract_multirow_section(ws, 11, "AC", "AD", {
        "well_id": "AD", "status": "AE", "flowing_hrs": "AF",
        "oil_prod": "AG", "gas_prod": "AH", "water_prod": "AI",
        "choke": "AJ", "whp": "AK", "wht": "AL", "dht": "AM",
        "dhp": "AN", "ap": "AO", "gor": "AP", "wlr": "AQ", "glr": "AR",
        "water_cut": "AS"
    })

    # Water Injection Wells (AU-BE, multiple per date)
    data["wi_wells"] = extract_multirow_section(ws, 11, "AU", "AV", {
        "well_id": "AV", "status": "AW", "flowing_hrs": "AX",
        "wi": "AY", "choke": "AZ", "whp": "BA", "wht": "BB",
        "fl_t": "BC", "fl_p": "BD", "ap": "BE"
    })

    # Daily Crude in Stock (BG-BO)
    data["stocks"] = extract_simple_section(ws, 11, "BG", {
        "crude_today": "BH", "cargo": "BI", "fpso_storage": "BJ",
        "fso_storage": "BK", "dead": "BL", "pumpable": "BM",
        "days_full": "BN", "total_field": "BO"
    })

    # Safety Report (BQ-BY)
    data["safety"] = extract_simple_section(ws, 11, "BQ", {
        "days_lti": "BR", "lti": "BS", "mti": "BT", "oil_spill": "BU",
        "gas_leak": "BV", "auth_notif": "BW", "quality_events": "BX", "drill": "BY"
    })

    # Chemicals (CA-CL, label rows)
    data["chemicals"] = extract_multirow_section(ws, 11, "CA", "CB", {
        "label": "CB", "demulsifier": "CC", "antifoam": "CD", "scale_inh": "CE",
        "deoiler": "CF", "biocide1": "CG", "biocide2": "CH", "sw_scale": "CI",
        "wax_inh": "CJ", "port_meoh": "CK", "stbd_meoh": "CL"
    })

    # Offtake Report (CN-CV)
    data["offtake"] = extract_simple_section(ws, 11, "CN", {
        "gross": "CO", "net": "CP", "bsw": "CQ", "density": "CR",
        "salt": "CS", "rvp": "CT", "temp": "CU", "press": "CV"
    })

    # Energy Consumption Report (CX-DC, label rows)
    data["energy"] = extract_multirow_section(ws, 11, "CX", "CY", {
        "label": "CY", "diesel_cons": "CZ", "diesel_stock": "DA",
        "fw_cons": "DB", "fw_stock": "DC"
    })

    return data


def extract_pci11(ws):
    """PETROCI CI-11"""
    data = {}

    # OIL PRODUCTION (E-M)
    data["oil"] = extract_simple_section(ws, 11, "E", {
        "gross": "F", "net": "G", "bsw": "H", "stocks": "I",
        "gross_export": "J", "net_export": "K", "water_overboard": "L",
        "oil_in_water": "M"
    })

    # GAS PRODUCTION AND EXPORT (O-W)
    data["gas"] = extract_simple_section(ws, 11, "O", {
        "lion_gas": "P", "panth_gas": "Q", "total_gas": "R",
        "gas_export": "S", "fuel_gas": "T", "flare": "U",
        "sys_sales": "V", "total_sales": "W"
    })

    # LION PRODUCTION WELLS (Y-AI, multiple per date)
    data["lion_wells"] = extract_multirow_section(ws, 11, "Y", "Z", {
        "well_id": "Z", "prod_sys": "AA", "ftp": "AB", "cp": "AC",
        "sitp": "AD", "choke": "AE", "hrs": "AF", "oil": "AG",
        "gas": "AH", "water": "AI"
    })

    # PANTHERE PRODUCTION WELLS (AK-AU, multiple per date)
    data["panth_wells"] = extract_multirow_section(ws, 11, "AK", "AL", {
        "well_id": "AL", "prod_sys": "AM", "ftp": "AN", "cp": "AO",
        "sitp": "AP", "choke": "AQ", "hrs": "AR", "oil": "AS",
        "gas": "AT", "water": "AU"
    })

    return data


def extract_fox(ws):
    """FOXTROT"""
    data = {}

    # Production et transfert HCL (E-M, label rows: Production / Transfert)
    data["hcl"] = extract_multirow_section(ws, 11, "E", "F", {
        "label": "F", "pfa": "G", "pfb": "H", "taboth": "I",
        "vridi_est": "J", "azito": "K", "vridi_ouest": "L", "total": "M"
    })

    # Volume Stock HCL (O-S)
    data["hcl_stock"] = extract_simple_section(ws, 11, "O", {
        "taboth": "P", "vridi_est": "Q", "azito": "R", "total": "S"
    })

    # Ventes journalières de Gaz (U-AF)
    data["gas_sales"] = extract_simple_section(ws, 11, "U", {
        "addah": "V", "azito123": "W", "azito4": "X", "azito1234": "Y",
        "petro_azito": "Z", "ciprel": "AA", "petro_vridi": "AB",
        "sir": "AC", "atinkou": "AD", "total_gas": "AE", "hcl_sir": "AF"
    })

    # Production journalière Eau (AH-AN)
    data["water"] = extract_simple_section(ws, 11, "AH", {
        "pfa": "AI", "pfb": "AJ", "taboth": "AK", "azito": "AL",
        "vridi_est": "AM", "total": "AN"
    })

    # Productions moyennes journalières de Gaz (AP-AY, label rows: PFA/PFB/Total)
    data["gas_avg"] = extract_multirow_section(ws, 11, "AP", "AQ", {
        "label": "AQ", "vol_export": "AR", "flare": "AS", "fuel_comp": "AT",
        "other_fuel": "AU", "total_prod": "AV", "gor": "AW", "wgr": "AX", "wor": "AY"
    })

    # PFA Production Wells (BA-BS, multiple per date)
    data["pfa_wells"] = extract_multirow_section(ws, 11, "BA", "BB", {
        "well_id": "BB", "gas": "BC", "hcl": "BD", "eau": "BE",
        "temps_f": "BF", "pfd": "BG", "tfd": "BH", "ptw": "BI", "ttw": "BJ",
        "ap_a": "BK", "ap_b": "BL", "ap_c": "BM", "p_flowline": "BN",
        "dp_duse": "BO", "dp_plate": "BP", "duse_act": "BQ", "duse_max": "BR", "sep": "BS"
    })

    # PFB Production Wells (BU-CM, multiple per date)
    data["pfb_wells"] = extract_multirow_section(ws, 11, "BU", "BV", {
        "well_id": "BV", "gas": "BW", "hcl": "BX", "eau": "BY",
        "temps_f": "BZ", "pfd": "CA", "tfd": "CB", "ptw": "CC", "ttw": "CD",
        "ap_a": "CE", "ap_b": "CF", "ap_c": "CG", "p_flowline": "CH",
        "dp_duse": "CI", "dp_plate": "CJ", "duse_act": "CK", "duse_max": "CL", "sep": "CM"
    })

    return data


def extract_esp(ws):
    """ESPOIR"""
    data = {}

    # Field Daily Oil Production (E-L)
    data["oil"] = extract_simple_section(ws, 11, "E", {
        "east": "F", "west": "G", "api": "H", "bsw": "I",
        "rvp": "J", "total": "K", "delta": "L"
    })

    # Field Daily Gas Production + Gas Lift + Export + Fuel (N-AI)
    data["gas"] = extract_simple_section(ws, 11, "N", {
        "east": "O", "west": "P",
        "glift_e": "Q", "glift_w": "R", "glift_dur": "S", "glift_pres": "T", "glift_temp": "U",
        "export_fpso": "V", "export_terminal": "W", "export_east": "X", "export_west": "Y",
        "export_pres": "Z", "export_temp": "AA", "dewpoint": "AB", "moisture": "AC",
        "compressors": "AD", "boilers": "AE", "total_flare": "AF",
        "total_fuel": "AG", "total_prod": "AH", "delta": "AI"
    })

    # Water Production + Water Injection (AK-AY)
    data["water"] = extract_simple_section(ws, 11, "AK", {
        "east_prod": "AL", "west_prod": "AM", "total_prod": "AN",
        "ulage": "AO", "overboard": "AP", "oil_in_water": "AQ",
        "east_wi": "AR", "west_wi": "AS", "total_wi": "AT", "wi_pres": "AU",
        "east_o2": "AV", "west_o2": "AW", "east_sulphite": "AX", "west_sulphite": "AY"
    })

    # East Production Wells (BA-BP, multiple per date)
    data["east_wells"] = extract_multirow_section(ws, 11, "BA", "BB", {
        "well_id": "BB", "oil": "BC", "gas": "BD", "gaslift": "BE",
        "gl_time": "BF", "water": "BG", "bhp": "BH", "bht": "BI",
        "thp": "BJ", "tht": "BK", "full_rate": "BL",
        "gor": "BM", "wlr": "BN", "glr": "BO", "water_cut": "BP"
    })

    # West Production Wells (BR-CG, multiple per date)
    data["west_wells"] = extract_multirow_section(ws, 11, "BR", "BS", {
        "well_id": "BS", "oil": "BT", "gas": "BU", "gaslift": "BV",
        "gl_time": "BW", "water": "BX", "bhp": "BY", "bht": "BZ",
        "thp": "CA", "tht": "CB", "full_rate": "CC",
        "gor": "CD", "wlr": "CE", "glr": "CF", "water_cut": "CG"
    })

    # Water Injection Wells (CI-CP, multiple per date)
    data["wi_wells"] = extract_multirow_section(ws, 11, "CI", "CJ", {
        "well_id": "CJ", "inj_annulus": "CK", "ann_pres": "CL",
        "ann_full_rate": "CM", "inj_tubing": "CN", "tub_pres": "CO",
        "tub_full_rate": "CP"
    })

    # Daily Crude in Stock (CR-CV)
    data["stocks"] = extract_simple_section(ws, 11, "CR", {
        "tov": "CS", "gov": "CT", "free_water": "CU", "net_oil": "CV"
    })

    # Safety Report (CX-CZ)
    data["safety"] = extract_simple_section(ws, 11, "CX", {
        "days_lti": "CY", "lti": "CZ"
    })

    # Chemical (DB-DR, label rows)
    data["chemicals"] = extract_multirow_section(ws, 11, "DB", "DC", {
        "label": "DC", "corro_e": "DD", "corro_w": "DE",
        "demusl_e": "DF", "demusl_w": "DG", "demusl_fpso": "DH",
        "wax_inh": "DI", "si4140": "DJ", "oxy_w": "DK", "oxy_e": "DL",
        "biocide": "DM", "antifoam": "DN", "h2s": "DO", "sand": "DP",
        "glycol": "DQ", "clarifier": "DR"
    })

    return data


# ─────────────────────────────────────────────
# MAIN EXTRACTION
# ─────────────────────────────────────────────

def extract_all(excel_path):
    """Lit le fichier Excel et retourne le dictionnaire RAW complet."""
    print(f"[{datetime.now():%H:%M:%S}] Chargement de {os.path.basename(excel_path)}...")
    wb = load_workbook(excel_path, read_only=True, data_only=True)

    raw = {}
    extractors = {
        "BP1":   ("BALEINE PHASE 1",  extract_bp1),
        "BP2":   ("BALEINE PHASE 2",  extract_bp2),
        "PCI11": ("PETROCI CI-11",    extract_pci11),
        "FOX":   ("FOXTROT",          extract_fox),
        "ESP":   ("ESPOIR",           extract_esp),
    }

    for code, (sheet_name, fn) in extractors.items():
        try:
            ws = wb[sheet_name]
            print(f"  Extraction {sheet_name}...", end=" ", flush=True)
            data = fn(ws)
            raw[code] = data
            # Stats
            total = sum(len(v) for v in data.values() if isinstance(v, list))
            print(f"{total} enregistrements")
        except Exception as e:
            print(f"ERREUR: {e}")
            raw[code] = {}

    wb.close()
    return raw


# ─────────────────────────────────────────────
# MISE A JOUR DU DASHBOARD
# ─────────────────────────────────────────────

def update_dashboard(excel_path, html_path):
    """Extrait les donnees et met a jour le fichier HTML."""
    try:
        raw = extract_all(excel_path)
    except Exception as e:
        print(f"[ERREUR] Impossible de lire le fichier Excel: {e}")
        return False

    # Serialisation JSON compacte
    json_str = json.dumps(raw, ensure_ascii=False, separators=(",", ":"))

    # Lecture du dashboard
    try:
        with open(html_path, "r", encoding="utf-8") as f:
            html = f.read()
    except Exception as e:
        print(f"[ERREUR] Impossible de lire le dashboard: {e}")
        return False

    # Remplacement de const RAW = {...};
    pattern = r"const RAW\s*=\s*\{.*?\};"
    new_line = f"const RAW = {json_str};"

    if not re.search(pattern, html, re.DOTALL):
        print("[ERREUR] Marqueur 'const RAW = {...};' introuvable dans le HTML.")
        return False

    # Utiliser une lambda pour eviter l'interpretation des backslashes par re.sub
    new_html = re.sub(pattern, lambda m: new_line, html, count=1, flags=re.DOTALL)

    # Sauvegarde avec horodatage de mise a jour
    ts_pattern = r'(id="lastUpdate"[^>]*>)[^<]*(</span>)'
    ts_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    new_html = re.sub(ts_pattern, rf"\g<1>{ts_str}\g<2>", new_html)

    # Ecriture
    try:
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(new_html)
        size_kb = os.path.getsize(html_path) // 1024
        print(f"[{datetime.now():%H:%M:%S}] Dashboard mis a jour ({size_kb} KB): {html_path}")
        return True
    except Exception as e:
        print(f"[ERREUR] Ecriture du dashboard: {e}")
        return False


# ─────────────────────────────────────────────
# MODE SURVEILLANCE (watchdog)
# ─────────────────────────────────────────────

def watch_mode(excel_path, html_path, debounce=5):
    """Surveille le fichier Excel et met a jour le dashboard automatiquement."""
    try:
        from watchdog.observers import Observer
        from watchdog.events import FileSystemEventHandler
    except ImportError:
        print("[ERREUR] watchdog non installe. Executez: pip install watchdog")
        sys.exit(1)

    print(f"[SURVEILLANCE] En attente de modifications sur:")
    print(f"  {excel_path}")
    print(f"Appuyez sur Ctrl+C pour arreter.\n")

    # Mise a jour initiale
    update_dashboard(excel_path, html_path)

    last_update = [0]

    class ExcelHandler(FileSystemEventHandler):
        def on_modified(self, event):
            if not event.is_directory and os.path.normcase(os.path.abspath(event.src_path)) == \
               os.path.normcase(os.path.abspath(excel_path)):
                now = time.time()
                if now - last_update[0] > debounce:
                    last_update[0] = now
                    print(f"\n[{datetime.now():%H:%M:%S}] Fichier Excel modifie, mise a jour...")
                    time.sleep(1)  # Attendre que le fichier soit libere
                    update_dashboard(excel_path, html_path)

    watch_dir = os.path.dirname(os.path.abspath(excel_path))
    observer = Observer()
    observer.schedule(ExcelHandler(), watch_dir, recursive=False)
    observer.start()

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
        print("\n[ARRET] Surveillance terminee.")
    observer.join()


# ─────────────────────────────────────────────
# POINT D'ENTREE
# ─────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Met a jour le dashboard Ivoirian Basin depuis le fichier Excel"
    )
    parser.add_argument(
        "--watch", "-w", action="store_true",
        help="Mode surveillance: met a jour automatiquement lors des modifications du fichier Excel"
    )
    parser.add_argument(
        "--excel", default=EXCEL_PATH,
        help="Chemin vers le fichier Excel (defaut: configuration interne)"
    )
    parser.add_argument(
        "--html", default=DASHBOARD_PATH,
        help="Chemin vers le fichier HTML dashboard (defaut: configuration interne)"
    )
    args = parser.parse_args()

    if args.watch:
        watch_mode(args.excel, args.html)
    else:
        success = update_dashboard(args.excel, args.html)
        sys.exit(0 if success else 1)
